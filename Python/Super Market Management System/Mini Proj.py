import openpyxl
wb1 = openpyxl.load_workbook("D:\\Admin.xlsx")
sh1 = ['Admin']
sh1 = wb1.active
wb2 = openpyxl.load_workbook("D:\\Sales.xlsx")
sh2 = ['Sales']
sh2 = wb2.active
wb3 = openpyxl.load_workbook("D:\\Purchase.xlsx")
sh3 = ['Purchase']
sh3 = wb3.active
wb4 = openpyxl.load_workbook("D:\\Employees.xlsx")
sh4 = ['Employees']
sh4 = wb4.active

def bill_menu():
    print("\n\n\t\t\tBIL MENU")
    print("1. Sale\n2. Purchase")
    ch = int(input("Please Enter your choise :...."))
    if ch == 1:
        sale = list()
        s_name = input("Enter Customer Name : ") 
        s_add = input("Enter Customer Address :")
        s_no = int(input("Enter Customer Mobile No. : "))
        a = 1
        while a == 1:
            s_code = int(input("Enter product code"))
            row = sh1.max_row
            col = sh1.max_column
            for i in range(1, row + 1):
                b =  sh1.cell(i, 1).value
                if s_code == b:
                    si_name = sh1.cell(i, 0).value
                    s_price = sh1.cell(i, 3).value
                    sa_quantity = sh1.cell(i, 2).value
                    qty = int(input("Enter Quantity"))
                    sh1.cell(i, 0).value = sh1.cell(i, 0).value - qty
                    amt = qty * sh1.cell(i, 3).value
                    tamt = tamt + amt
                    sale = [bill_no, s_name, s_add, s_no, si_name, b, s_price, qty, tamt]
                    for j in sale:
                        sh2.append(i)
                    ch = int(input("1. Yes\n2. NO\nPlease Enter your choise :...."))
                    if ch2 == 1:
                        a = 1
                    else :
                        a = 0
                        break



def admin_menu():
    print("\n\n\t\t\t ADMIN MENU  ");
    print("\n1. Create Product \n2. Display all Product \n3. Modify Product \n4. Delete Product \n5. Back to Main Menu")
    ch = int(input("Please Enter your choise :...."))
    a = 1
# Create New Product
    if ch == 1:
        while a == 1:
            print("Do you want to ADD more Products\n\n [1. YES / 2. NO] : "); 
            ch1 = int(input("Please Enter your choise :...."))
            if ch1 == 1:
                pro =list()
                p_name = input("Enter Product Name : ")
                p_hsn = ch1 = int(input("Enter HSN Code : "))
                p_quantity = int(input("Enter Product Quantity : "))                
                p_price = int(input("Enter Product Price : "))
                tup = (p_name, p_hsn, p_quantity, p_price)
                print(tup)
                pro.append(tup)
                print(pro)
                for i in pro:
                    sh1.append(i)
                wb1.save("D:\\Admin.xlsx")
            elif ch1 == 2:
                admin_menu()
            else:
                menu()

# Print Product List
    elif ch == 2:
        row = sh1.max_row
        col = sh1.max_column
        for i in range (1, row + 1):
            for j in range (1, col + 1):
                print(sh1.cell(i, j).value, end="\t\t\t")
            print("\n")
        wb1.save("D:\\Admin.xlsx")
        admin_menu()

# Modify Product Details
    elif ch == 3:
        row = sh1.max_row
        col = sh1.max_column
        b = int(input("HSN code of the product to be modified : "))
        for i in range (1, row + 1):
            for j in range (1, col + 1):
                c = sh1.cell(i, j).value
                if b == c:
                    p_name = input("Enter Modified Product Name : ")
                    sh1.cell(row = i, column = 1).value = p_name
                    p_hsn = ch1 = int(input("Enter Modified HSN Code : "))
                    sh1.cell(row = i, column = 2).value = p_hsn
                    p_quantity = int(input("Enter Modified Product Quantity : "))
                    sh1.cell(row = i, column = 3).value = p_quantity
                    p_price = int(input("Enter Modified Product Price : "))
                    sh1.cell(row = i, column = 4).value = p_price
                    break
        wb1.save("D:\\Admin.xlsx")
        admin_menu()

# Delete Product Details
    elif ch == 4:
        row = sh1.max_row
        col = sh1.max_column
        b = int(input("HSN code of the product to be modified : "))
        for i in range (1, row + 1):
            c = sh1.cell(i, 1).value
            if b == c:
                sh1.delete_rows(i, 1)
                wb1.save("D:\\Admin.xlsx")
                break
        admin_menu()


    elif ch == 5:
        wb1.save("D:\\Admin.xlsx")
        menu()    



def emp_menu():
    print("\n\n\t\t\t EMPLOYEE MENU  ");
    print("\n1. New Employee \n2. Display all Employee \n3. Modify Employee \n4. Delete Employee \n5. Back to Main Menu")
    ch = int(input("Please Enter your choise :...."))
    a = 1
# Add New Employee
    if ch == 1:
        while a == 1:
            print("Do you want to ADD more Products\n\n [1. YES / 2. NO] : "); 
            ch1 = int(input("Please Enter your choise :...."))
            if ch1 == 1:
                emp =list()
                e_id = int(input("Enter Employee ID : "))
                e_name =  input("Enter Employee Name : ")
                e_no = int(input("Enter Employee Mob. No. : "))                
                e_add = input("Enter Employee Address : ")
                tup1 = (e_id, e_name, e_no, e_add)
                emp.append(tup1)
                for k in emp:
                    sh4.append(i)
                wb4.save("D:\\Employees.xlsx")
            elif ch1 == 2:
                emp_menu()
            else:
                menu()

# Print Employee List             
    elif ch == 2:
        row = sh4.max_row
        col = sh4.max_column
        for i in range (1, row + 1):
            for j in range (0, column + 1):
                print(sh4.cell(i, j).value, end = "\t")
        wb4.save("D:\\Employees.xlsx")
        emp_menu()
        
# Modify Employee Details
    elif ch == 3:
        b = int(input("Enter ID od an Employee to be modified : "))
        for i in range (1, row + 1):
            for j in range (0, column + 1):
                c = sh4.cell(i, j).value
                if b == c:
                    e_id = int(input("Enter Employee ID : "))
                    sh4.cell(row = i, column = j).value = e_id
                    e_name =  input("Enter Employee Name : ")
                    sh4.cell(row = i, column = j).value = e_name
                    e_no = int(input("Enter Employee Mob. No. : "))                
                    sh4.cell(row = i, column = j).value = e_no
                    e_add = input("Enter Employee Address : ")
                    sh4.cell(row = i, column = j).value = e_add
                    break
        wb4.save("D:\\Employees.xlsx")
        emp_menu()   
            
# Delete Empoyee Details
    elif ch == 4:
        b = int(input("Enter ID od an Employee to be Deleted : "))
        for i in range (1, row + 1):
            c = sh4.cell(i, 0).value
            if b == c:
                sh4.delete_rows(i, 1)
                wb4.save("D:\\Employees.xlsx")
                break
        emp_menu()


    elif ch == 5:
        wb4.save("D:\\Employees.xlsx")
        menu()    




                
def menu():
    print("\t\t\tMAIN MENU")
    print("1. Billing")
    print("2. Administration")
    print("3. Employees")
    print("4. Exit")
    ch = int(input("Please Enter your choise :...."))
    if ch == 1:
        bill_menu()
    elif ch == 2:
        admin_menu()
        
    elif ch == 3:
        emp_menu()
        
    elif ch == 4:
        exit()


print("______________________________________________________________")
print("\t\tSUPER MARKET MANAGEMENT SYSTEM")
print("______________________________________________________________")
menu()
